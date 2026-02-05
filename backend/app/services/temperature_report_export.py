"""
Temperature Report Export Service
온도 분석 보고서 엑셀 생성 서비스
Phase 3-A Part 5: 고급 분석 대시보드
"""
import io
from datetime import datetime, timedelta
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from sqlalchemy.orm import Session

from app.services.temperature_analytics import TemperatureAnalytics


class TemperatureReportExporter:
    """온도 분석 보고서 엑셀 생성"""
    
    def __init__(self, db: Session):
        self.db = db
        self.analytics = TemperatureAnalytics(db)
    
    def generate_compliance_report(
        self,
        start_date: datetime,
        end_date: datetime,
        vehicle_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        준수 보고서 엑셀 생성
        
        Args:
            start_date: 시작 날짜
            end_date: 종료 날짜
            vehicle_id: 차량 ID (optional)
            
        Returns:
            엑셀 파일 (BytesIO)
        """
        wb = Workbook()
        
        # 1. 요약 시트
        ws_summary = wb.active
        ws_summary.title = "요약"
        self._create_summary_sheet(ws_summary, start_date, end_date, vehicle_id)
        
        # 2. 위반 내역 시트
        ws_violations = wb.create_sheet("위반 내역")
        self._create_violations_sheet(ws_violations, start_date, end_date, vehicle_id)
        
        # 3. 차량별 통계 시트
        if not vehicle_id:
            ws_vehicles = wb.create_sheet("차량별 통계")
            self._create_vehicles_stats_sheet(ws_vehicles, start_date, end_date)
        
        # 엑셀 파일을 BytesIO로 저장
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _create_summary_sheet(
        self,
        ws,
        start_date: datetime,
        end_date: datetime,
        vehicle_id: Optional[int]
    ):
        """요약 시트 생성"""
        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        
        # 제목
        ws["A1"] = "온도 준수 보고서"
        ws["A1"].font = Font(size=18, bold=True)
        ws.merge_cells("A1:D1")
        
        # 기본 정보
        ws["A3"] = "보고 기간:"
        ws["B3"] = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        
        ws["A4"] = "생성 일시:"
        ws["B4"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 준수 보고서 가져오기
        report = self.analytics.get_compliance_report(start_date, end_date, vehicle_id)
        
        # 핵심 지표
        ws["A6"] = "핵심 지표"
        ws["A6"].font = header_font
        ws["A6"].fill = header_fill
        ws.merge_cells("A6:B6")
        
        metrics = [
            ("전체 기록 수", report["total_records"]),
            ("준수 기록 수", report["compliant_records"]),
            ("위반 기록 수", report["violation_records"]),
            ("준수율 (%)", f"{report['compliance_rate']}%")
        ]
        
        row = 7
        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        
        # 위반 요약
        if report["violation_summary"]["by_type"]:
            ws["A12"] = "위반 유형별 통계"
            ws["A12"].font = header_font
            ws["A12"].fill = header_fill
            ws.merge_cells("A12:B12")
            
            row = 13
            for v_type, count in report["violation_summary"]["by_type"].items():
                ws[f"A{row}"] = v_type
                ws[f"B{row}"] = count
                row += 1
    
    def _create_violations_sheet(
        self,
        ws,
        start_date: datetime,
        end_date: datetime,
        vehicle_id: Optional[int]
    ):
        """위반 내역 시트 생성"""
        # 헤더
        headers = ["시각", "차량 번호", "센서", "온도 (°C)", "위반 유형", "위도", "경도"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 위반 데이터
        report = self.analytics.get_compliance_report(start_date, end_date, vehicle_id)
        violations = report["violations"]
        
        for row_idx, violation in enumerate(violations, 2):
            ws.cell(row=row_idx, column=1, value=violation["timestamp"])
            ws.cell(row=row_idx, column=2, value=violation["vehicle_number"])
            ws.cell(row=row_idx, column=3, value=violation["sensor"])
            ws.cell(row=row_idx, column=4, value=violation["temperature"])
            ws.cell(row=row_idx, column=5, value=violation["violation_type"])
            ws.cell(row=row_idx, column=6, value=violation.get("latitude", ""))
            ws.cell(row=row_idx, column=7, value=violation.get("longitude", ""))
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
    
    def _create_vehicles_stats_sheet(
        self,
        ws,
        start_date: datetime,
        end_date: datetime
    ):
        """차량별 통계 시트 생성"""
        from app.models.vehicle import Vehicle
        
        # 헤더
        headers = ["차량 번호", "성능 점수", "등급", "준수율 A (%)", "준수율 B (%)", "안정성 A", "안정성 B"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 차량별 데이터
        vehicles = self.db.query(Vehicle).all()
        days = (end_date - start_date).days
        
        for row_idx, vehicle in enumerate(vehicles, 2):
            try:
                performance = self.analytics.get_vehicle_performance_score(vehicle.id, days)
                
                ws.cell(row=row_idx, column=1, value=vehicle.plate_number)
                ws.cell(row=row_idx, column=2, value=performance["score"])
                ws.cell(row=row_idx, column=3, value=performance["grade"])
                ws.cell(row=row_idx, column=4, value=performance["metrics"]["sensor_a"]["compliance_rate"])
                ws.cell(row=row_idx, column=5, value=performance["metrics"]["sensor_b"]["compliance_rate"])
                ws.cell(row=row_idx, column=6, value=performance["metrics"]["sensor_a"]["stability"])
                ws.cell(row=row_idx, column=7, value=performance["metrics"]["sensor_b"]["stability"])
            except Exception as e:
                ws.cell(row=row_idx, column=1, value=vehicle.plate_number)
                ws.cell(row=row_idx, column=2, value="N/A")
        
        # 열 너비 조정
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def generate_performance_report(
        self,
        days: int = 30
    ) -> io.BytesIO:
        """
        차량 성능 보고서 엑셀 생성
        
        Args:
            days: 분석 기간 (일)
            
        Returns:
            엑셀 파일 (BytesIO)
        """
        from app.models.vehicle import Vehicle
        
        wb = Workbook()
        ws = wb.active
        ws.title = "차량 성능 순위"
        
        # 제목
        ws["A1"] = f"차량 온도 관리 성능 보고서 ({days}일)"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:H1")
        
        # 헤더
        headers = ["순위", "차량 번호", "점수", "등급", "준수율 (%)", "안정성", "데이터 수집률 (%)", "권장사항"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # 차량 성능 데이터
        vehicles = self.db.query(Vehicle).all()
        performances = []
        
        for vehicle in vehicles:
            try:
                perf = self.analytics.get_vehicle_performance_score(vehicle.id, days)
                performances.append(perf)
            except:
                continue
        
        # 점수 기준 정렬
        performances.sort(key=lambda x: x["score"], reverse=True)
        
        # 데이터 입력
        for rank, perf in enumerate(performances, 1):
            row = rank + 3
            
            avg_compliance = (
                perf["metrics"]["sensor_a"]["compliance_rate"] + 
                perf["metrics"]["sensor_b"]["compliance_rate"]
            ) / 2
            
            avg_stability = (
                perf["metrics"]["sensor_a"]["stability"] + 
                perf["metrics"]["sensor_b"]["stability"]
            ) / 2
            
            recommendations = "; ".join(perf["recommendations"][:2])  # 처음 2개만
            
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=perf["vehicle_number"])
            ws.cell(row=row, column=3, value=perf["score"])
            ws.cell(row=row, column=4, value=perf["grade"])
            ws.cell(row=row, column=5, value=round(avg_compliance, 2))
            ws.cell(row=row, column=6, value=round(avg_stability, 2))
            ws.cell(row=row, column=7, value=perf["metrics"]["data_collection_rate"])
            ws.cell(row=row, column=8, value=recommendations)
            
            # 점수에 따른 색상
            score_cell = ws.cell(row=row, column=3)
            if perf["score"] >= 90:
                score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif perf["score"] >= 70:
                score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif perf["score"] < 60:
                score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 50
        
        # 엑셀 파일을 BytesIO로 저장
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
