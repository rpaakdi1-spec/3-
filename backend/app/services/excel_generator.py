"""
Excel 생성 서비스
OpenPyXL을 사용하여 Excel 파일 생성
"""

from io import BytesIO
from typing import Dict, Any, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Excel 리포트 생성기"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        self.normal_font = Font(name="맑은 고딕", size=10)
        self.title_font = Font(name="맑은 고딕", size=14, bold=True)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_header_style(self, sheet, row_num: int, col_start: int, col_end: int):
        """
        헤더 스타일 적용
        """
        for col in range(col_start, col_end + 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
    
    def _apply_data_style(self, sheet, row_start: int, row_end: int, col_start: int, col_end: int):
        """
        데이터 셀 스타일 적용
        """
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = self.normal_font
                cell.border = self.thin_border
    
    def _format_currency(self, value: float) -> str:
        """
        통화 포맷 (원화)
        """
        return f"₩{value:,.0f}"
    
    def _format_percentage(self, value: float) -> str:
        """
        퍼센트 포맷
        """
        return f"{value:.1f}%"
    
    def generate_financial_dashboard_excel(
        self,
        summary: Dict[str, Any],
        monthly_trends: List[Dict[str, Any]],
        top_clients: List[Dict[str, Any]],
        start_date: str,
        end_date: str
    ) -> bytes:
        """
        재무 대시보드 Excel 생성
        
        Args:
            summary: 재무 요약 (14개 지표)
            monthly_trends: 월별 추이 데이터
            top_clients: Top 10 고객 목록
            start_date: 시작일 (YYYY-MM-DD)
            end_date: 종료일 (YYYY-MM-DD)
        
        Returns:
            Excel 바이너리 데이터
        """
        wb = Workbook()
        
        # Sheet 1: 요약 (Summary)
        ws_summary = wb.active
        ws_summary.title = "요약"
        self._create_summary_sheet(ws_summary, summary, start_date, end_date)
        
        # Sheet 2: 월별 데이터 (Monthly Data)
        ws_monthly = wb.create_sheet("월별 데이터")
        self._create_monthly_data_sheet(ws_monthly, monthly_trends)
        
        # Sheet 3: Top 고객 (Top Clients)
        ws_clients = wb.create_sheet("Top 고객")
        self._create_top_clients_sheet(ws_clients, top_clients)
        
        # Sheet 4: 차트 (Charts)
        ws_chart = wb.create_sheet("차트")
        self._create_chart_sheet(ws_chart, monthly_trends)
        
        # Excel 파일을 BytesIO에 저장
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.read()
    
    def _create_summary_sheet(self, sheet, summary: Dict[str, Any], start_date: str, end_date: str):
        """
        요약 시트 생성
        """
        # 제목
        sheet['A1'] = "재무 대시보드 리포트"
        sheet['A1'].font = self.title_font
        
        sheet['A2'] = f"기간: {start_date} ~ {end_date}"
        sheet['A2'].font = self.normal_font
        
        sheet['A3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A3'].font = self.normal_font
        
        # 헤더
        row = 5
        sheet[f'A{row}'] = "지표"
        sheet[f'B{row}'] = "값"
        self._apply_header_style(sheet, row, 1, 2)
        
        # 데이터
        metrics = [
            ("총 수익", summary.get("total_revenue", 0), "currency"),
            ("청구액", summary.get("total_invoiced", 0), "currency"),
            ("수금액", summary.get("total_paid", 0), "currency"),
            ("미수금", summary.get("total_outstanding", 0), "currency"),
            ("수금률", summary.get("payment_rate", 0), "percentage"),
            ("연체 건수", summary.get("overdue_count", 0), "number"),
            ("연체 금액", summary.get("overdue_amount", 0), "currency"),
            ("정산 대기 금액", summary.get("pending_settlement_amount", 0), "currency"),
            ("현금 유입", summary.get("cash_in", 0), "currency"),
            ("현금 유출", summary.get("cash_out", 0), "currency"),
            ("순 현금 흐름", summary.get("net_cash_flow", 0), "currency"),
        ]
        
        for idx, (metric_name, value, fmt) in enumerate(metrics, start=row+1):
            sheet[f'A{idx}'] = metric_name
            if fmt == "currency":
                sheet[f'B{idx}'] = self._format_currency(value)
            elif fmt == "percentage":
                sheet[f'B{idx}'] = self._format_percentage(value)
            else:
                sheet[f'B{idx}'] = value
        
        self._apply_data_style(sheet, row+1, row+len(metrics), 1, 2)
        
        # 열 너비 조정
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
    
    def _create_monthly_data_sheet(self, sheet, monthly_trends: List[Dict[str, Any]]):
        """
        월별 데이터 시트 생성
        """
        # 헤더
        headers = ["월", "수익", "청구액", "수금액", "미수금", "수금률"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
        
        self._apply_header_style(sheet, 1, 1, len(headers))
        
        # 데이터
        for row, trend in enumerate(monthly_trends, start=2):
            sheet.cell(row=row, column=1).value = trend.get("month", "")
            sheet.cell(row=row, column=2).value = self._format_currency(trend.get("revenue", 0))
            sheet.cell(row=row, column=3).value = self._format_currency(trend.get("invoiced", 0))
            sheet.cell(row=row, column=4).value = self._format_currency(trend.get("paid", 0))
            sheet.cell(row=row, column=5).value = self._format_currency(trend.get("outstanding", 0))
            sheet.cell(row=row, column=6).value = self._format_percentage(trend.get("payment_rate", 0))
        
        if monthly_trends:
            self._apply_data_style(sheet, 2, 1 + len(monthly_trends), 1, len(headers))
        
        # 열 너비 조정
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_top_clients_sheet(self, sheet, top_clients: List[Dict[str, Any]]):
        """
        Top 고객 시트 생성
        """
        # 헤더
        headers = ["순위", "고객명", "총 매출", "청구액", "수금액", "미수금"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
        
        self._apply_header_style(sheet, 1, 1, len(headers))
        
        # 데이터
        for row, client in enumerate(top_clients, start=2):
            sheet.cell(row=row, column=1).value = row - 1  # 순위
            sheet.cell(row=row, column=2).value = client.get("client_name", "")
            sheet.cell(row=row, column=3).value = self._format_currency(client.get("total_revenue", 0))
            sheet.cell(row=row, column=4).value = self._format_currency(client.get("invoiced", 0))
            sheet.cell(row=row, column=5).value = self._format_currency(client.get("paid", 0))
            sheet.cell(row=row, column=6).value = self._format_currency(client.get("outstanding", 0))
        
        if top_clients:
            self._apply_data_style(sheet, 2, 1 + len(top_clients), 1, len(headers))
        
        # 열 너비 조정
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        sheet.column_dimensions['B'].width = 25  # 고객명 열 넓게
    
    def _create_chart_sheet(self, sheet, monthly_trends: List[Dict[str, Any]]):
        """
        차트 시트 생성
        """
        if not monthly_trends:
            sheet['A1'] = "차트 생성을 위한 데이터가 없습니다."
            return
        
        # 차트 데이터 삽입
        headers = ["월", "수익", "청구액", "수금액"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col).value = header
        
        for row, trend in enumerate(monthly_trends, start=2):
            sheet.cell(row=row, column=1).value = trend.get("month", "")
            sheet.cell(row=row, column=2).value = trend.get("revenue", 0)
            sheet.cell(row=row, column=3).value = trend.get("invoiced", 0)
            sheet.cell(row=row, column=4).value = trend.get("paid", 0)
        
        # 라인 차트 생성
        chart = LineChart()
        chart.title = "월별 수익 추이"
        chart.style = 13
        chart.x_axis.title = "월"
        chart.y_axis.title = "금액 (원)"
        
        data = Reference(sheet, min_col=2, min_row=1, max_row=len(monthly_trends) + 1, max_col=4)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(monthly_trends) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # 차트 삽입
        sheet.add_chart(chart, "F2")


# Singleton 인스턴스
excel_generator = ExcelGenerator()
