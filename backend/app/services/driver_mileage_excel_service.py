"""
Driver Mileage Excel Export Service
운전자 주행거리 Excel 다운로드 서비스
"""
from datetime import datetime, date, timedelta
from typing import List, Optional, Dict, Any
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from app.models.driver_daily_mileage import DriverDailyMileage
from loguru import logger


class DriverMileageExcelService:
    """운전자 주행거리 Excel 생성 서비스"""
    
    def __init__(self, db: Session):
        self.db = db
        
        # 스타일 정의
        self.title_font = Font(name='맑은 고딕', size=16, bold=True, color="1F4E78")
        self.header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.data_font = Font(name='맑은 고딕', size=10)
        self.data_alignment_center = Alignment(horizontal='center', vertical='center')
        self.data_alignment_right = Alignment(horizontal='right', vertical='center')
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 강조 색상
        self.highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    def _apply_header_style(self, ws, row_num: int, col_start: int, col_end: int):
        """헤더 스타일 적용"""
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
    
    def _apply_data_style(self, ws, row_start: int, row_end: int, col_start: int, col_end: int, 
                         align='center'):
        """데이터 스타일 적용"""
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.alignment = self.data_alignment_center if align == 'center' else self.data_alignment_right
                cell.border = self.border
    
    def _auto_adjust_column_width(self, ws):
        """컬럼 너비 자동 조정"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
    
    def generate_daily_report(
        self,
        target_date: date,
        driver_name: Optional[str] = None
    ) -> BytesIO:
        """일별 운전자 주행거리 리포트 생성"""
        logger.info(f"Generating daily driver mileage report for {target_date}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{target_date.strftime('%Y%m%d')}"
        
        # 제목
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"운전자별 주행거리 일일 리포트 - {target_date.strftime('%Y년 %m월 %d일')}"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 헤더
        headers = [
            '순위', '운전자명', '총 주행거리(km)', '주행시간(분)', 
            '엔진가동(분)', '공회전(분)', '평균속도(km/h)', 
            '최고속도(km/h)', 'GPS포인트', '차량수'
        ]
        
        ws.append([])  # 빈 행
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
        
        self._apply_header_style(ws, header_row, 1, len(headers))
        
        # 데이터 조회
        query = self.db.query(DriverDailyMileage).filter(
            DriverDailyMileage.date == target_date,
            DriverDailyMileage.calculation_method == 'vehicle_based'
        )
        
        if driver_name:
            query = query.filter(DriverDailyMileage.notes.like(f'%{driver_name}%'))
        
        results = query.order_by(DriverDailyMileage.total_distance_km.desc()).all()
        
        # 데이터 입력
        data_start_row = header_row + 1
        for rank, record in enumerate(results, 1):
            driver_name_val = record.notes.replace("차량기반:", "") if record.notes else "Unknown"
            
            row_data = [
                rank,
                driver_name_val,
                round(record.total_distance_km, 2),
                record.total_driving_minutes,
                record.engine_on_minutes,
                record.idle_minutes,
                round(record.avg_speed_kmh, 1) if record.avg_speed_kmh else 0,
                round(record.max_speed_kmh, 1) if record.max_speed_kmh else 0,
                record.gps_point_count,
                record.vehicle_count
            ]
            ws.append(row_data)
        
        data_end_row = data_start_row + len(results) - 1
        
        # 데이터 스타일 적용
        if len(results) > 0:
            self._apply_data_style(ws, data_start_row, data_end_row, 1, len(headers))
            
            # 숫자 포맷 적용
            for row in range(data_start_row, data_end_row + 1):
                ws.cell(row=row, column=3).number_format = '#,##0.00'  # 주행거리
                ws.cell(row=row, column=7).number_format = '0.0'       # 평균속도
                ws.cell(row=row, column=8).number_format = '0.0'       # 최고속도
        
        # 합계 행
        total_row = data_end_row + 2
        ws.cell(row=total_row, column=1).value = "합계"
        ws.cell(row=total_row, column=2).value = f"{len(results)}명"
        ws.cell(row=total_row, column=3).value = sum(r.total_distance_km for r in results)
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        
        # 합계 행 스타일
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = self.total_fill
            cell.font = Font(name='맑은 고딕', size=10, bold=True)
            cell.border = self.border
            cell.alignment = self.data_alignment_center
        
        # 컬럼 너비 조정
        self._auto_adjust_column_width(ws)
        
        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Daily report generated: {len(results)} drivers")
        return output
    
    def generate_weekly_report(
        self,
        end_date: date,
        driver_name: Optional[str] = None
    ) -> BytesIO:
        """주간 운전자 주행거리 리포트 생성 (최근 7일)"""
        start_date = end_date - timedelta(days=6)
        logger.info(f"Generating weekly driver mileage report from {start_date} to {end_date}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "주간 리포트"
        
        # 제목
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"운전자별 주행거리 주간 리포트 ({start_date} ~ {end_date})"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 헤더
        headers = [
            '순위', '운전자명', '총 주행거리(km)', '총 주행시간(h)', 
            '평균속도(km/h)', '최고속도(km/h)', '운행일수', '일평균 주행거리(km)'
        ]
        
        ws.append([])
        header_row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col).value = header
        
        self._apply_header_style(ws, header_row, 1, len(headers))
        
        # 데이터 조회 (주간 집계)
        query = self.db.query(
            DriverDailyMileage.notes,
            func.sum(DriverDailyMileage.total_distance_km).label('total_distance'),
            func.sum(DriverDailyMileage.total_driving_minutes).label('total_minutes'),
            func.avg(DriverDailyMileage.avg_speed_kmh).label('avg_speed'),
            func.max(DriverDailyMileage.max_speed_kmh).label('max_speed'),
            func.count(DriverDailyMileage.id).label('driving_days')
        ).filter(
            and_(
                DriverDailyMileage.date >= start_date,
                DriverDailyMileage.date <= end_date,
                DriverDailyMileage.calculation_method == 'vehicle_based'
            )
        ).group_by(DriverDailyMileage.notes)
        
        if driver_name:
            query = query.filter(DriverDailyMileage.notes.like(f'%{driver_name}%'))
        
        results = query.order_by(func.sum(DriverDailyMileage.total_distance_km).desc()).all()
        
        # 데이터 입력
        data_start_row = header_row + 1
        for rank, record in enumerate(results, 1):
            driver_name_val = record.notes.replace("차량기반:", "") if record.notes else "Unknown"
            total_hours = record.total_minutes / 60 if record.total_minutes else 0
            avg_daily_distance = record.total_distance / record.driving_days if record.driving_days else 0
            
            row_data = [
                rank,
                driver_name_val,
                round(record.total_distance, 2),
                round(total_hours, 1),
                round(record.avg_speed or 0, 1),
                round(record.max_speed or 0, 1),
                record.driving_days,
                round(avg_daily_distance, 2)
            ]
            ws.append(row_data)
        
        data_end_row = data_start_row + len(results) - 1
        
        # 스타일 적용
        if len(results) > 0:
            self._apply_data_style(ws, data_start_row, data_end_row, 1, len(headers))
            
            for row in range(data_start_row, data_end_row + 1):
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                ws.cell(row=row, column=4).number_format = '0.0'
                ws.cell(row=row, column=5).number_format = '0.0'
                ws.cell(row=row, column=6).number_format = '0.0'
                ws.cell(row=row, column=8).number_format = '#,##0.00'
        
        self._auto_adjust_column_width(ws)
        
        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Weekly report generated: {len(results)} drivers")
        return output
    
    def generate_monthly_report(
        self,
        year: int,
        month: int,
        driver_name: Optional[str] = None
    ) -> BytesIO:
        """월간 운전자 주행거리 리포트 생성"""
        from calendar import monthrange
        
        start_date = date(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        
        logger.info(f"Generating monthly driver mileage report for {year}-{month:02d}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}{month:02d}"
        
        # 제목
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"운전자별 주행거리 월간 리포트 - {year}년 {month}월"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 헤더
        headers = [
            '순위', '운전자명', '총 주행거리(km)', '총 주행시간(h)', 
            '총 공회전시간(h)', '평균속도(km/h)', '최고속도(km/h)', 
            '운행일수', '일평균 주행거리(km)', '사용 차량수'
        ]
        
        ws.append([])
        header_row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col).value = header
        
        self._apply_header_style(ws, header_row, 1, len(headers))
        
        # 데이터 조회 (월간 집계)
        query = self.db.query(
            DriverDailyMileage.notes,
            func.sum(DriverDailyMileage.total_distance_km).label('total_distance'),
            func.sum(DriverDailyMileage.total_driving_minutes).label('total_minutes'),
            func.sum(DriverDailyMileage.idle_minutes).label('total_idle'),
            func.avg(DriverDailyMileage.avg_speed_kmh).label('avg_speed'),
            func.max(DriverDailyMileage.max_speed_kmh).label('max_speed'),
            func.count(DriverDailyMileage.id).label('driving_days'),
            func.count(func.distinct(DriverDailyMileage.vehicle_ids)).label('vehicle_count')
        ).filter(
            and_(
                DriverDailyMileage.date >= start_date,
                DriverDailyMileage.date <= end_date,
                DriverDailyMileage.calculation_method == 'vehicle_based'
            )
        ).group_by(DriverDailyMileage.notes)
        
        if driver_name:
            query = query.filter(DriverDailyMileage.notes.like(f'%{driver_name}%'))
        
        results = query.order_by(func.sum(DriverDailyMileage.total_distance_km).desc()).all()
        
        # 데이터 입력
        data_start_row = header_row + 1
        for rank, record in enumerate(results, 1):
            driver_name_val = record.notes.replace("차량기반:", "") if record.notes else "Unknown"
            total_hours = record.total_minutes / 60 if record.total_minutes else 0
            idle_hours = record.total_idle / 60 if record.total_idle else 0
            avg_daily_distance = record.total_distance / record.driving_days if record.driving_days else 0
            
            row_data = [
                rank,
                driver_name_val,
                round(record.total_distance, 2),
                round(total_hours, 1),
                round(idle_hours, 1),
                round(record.avg_speed or 0, 1),
                round(record.max_speed or 0, 1),
                record.driving_days,
                round(avg_daily_distance, 2),
                record.vehicle_count
            ]
            ws.append(row_data)
        
        data_end_row = data_start_row + len(results) - 1
        
        # 스타일 적용
        if len(results) > 0:
            self._apply_data_style(ws, data_start_row, data_end_row, 1, len(headers))
            
            for row in range(data_start_row, data_end_row + 1):
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                ws.cell(row=row, column=4).number_format = '0.0'
                ws.cell(row=row, column=5).number_format = '0.0'
                ws.cell(row=row, column=6).number_format = '0.0'
                ws.cell(row=row, column=7).number_format = '0.0'
                ws.cell(row=row, column=9).number_format = '#,##0.00'
        
        # 합계 행
        total_row = data_end_row + 2
        ws.cell(row=total_row, column=1).value = "합계"
        ws.cell(row=total_row, column=2).value = f"{len(results)}명"
        ws.cell(row=total_row, column=3).value = sum(r.total_distance for r in results)
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = self.total_fill
            cell.font = Font(name='맑은 고딕', size=10, bold=True)
            cell.border = self.border
            cell.alignment = self.data_alignment_center
        
        self._auto_adjust_column_width(ws)
        
        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Monthly report generated: {len(results)} drivers")
        return output
    
    def generate_custom_report(
        self,
        records: List[DriverDailyMileage],
        title: str,
        period: str
    ) -> bytes:
        """
        커스텀 리포트 생성 (모바일용)
        
        Args:
            records: 주행거리 기록 리스트
            title: 리포트 제목
            period: 기간 표시 문자열
        
        Returns:
            bytes: Excel 파일 바이트
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "주행거리 리포트"
        
        # 제목
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 기간
        ws.merge_cells('A2:J2')
        period_cell = ws['A2']
        period_cell.value = f"조회 기간: {period}"
        period_cell.font = Font(name='맑은 고딕', size=11, color="666666")
        period_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20
        
        # 헤더
        headers = [
            "날짜", "주행거리(km)", "주행시간", "엔진가동",
            "공회전", "평균속도", "최고속도", "시작시간", "종료시간", "차량수"
        ]
        header_row = 4
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col_num).value = header
        
        self._apply_header_style(ws, header_row, 1, len(headers))
        
        # 데이터 행
        data_start_row = header_row + 1
        
        for record in records:
            driving_hours = (record.total_driving_minutes or 0) / 60
            engine_hours = (record.engine_on_minutes or 0) / 60
            idle_hours = (record.idle_minutes or 0) / 60
            
            row_data = [
                str(record.date),
                round(record.total_distance_km or 0, 2),
                round(driving_hours, 1),
                round(engine_hours, 1),
                round(idle_hours, 1),
                round(record.avg_speed_kmh or 0, 1),
                round(record.max_speed_kmh or 0, 1),
                record.start_time.strftime('%H:%M') if record.start_time else '-',
                record.end_time.strftime('%H:%M') if record.end_time else '-',
                record.vehicle_count or 0
            ]
            ws.append(row_data)
        
        data_end_row = data_start_row + len(records) - 1
        
        # 스타일 적용
        if len(records) > 0:
            self._apply_data_style(ws, data_start_row, data_end_row, 1, len(headers))
            
            for row in range(data_start_row, data_end_row + 1):
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                ws.cell(row=row, column=3).number_format = '0.0'
                ws.cell(row=row, column=4).number_format = '0.0'
                ws.cell(row=row, column=5).number_format = '0.0'
                ws.cell(row=row, column=6).number_format = '0.0'
                ws.cell(row=row, column=7).number_format = '0.0'
        
        # 합계 행
        total_row = data_end_row + 2
        total_distance = sum(r.total_distance_km or 0 for r in records)
        total_driving_minutes = sum(r.total_driving_minutes or 0 for r in records)
        
        ws.cell(row=total_row, column=1).value = "합계"
        ws.cell(row=total_row, column=2).value = round(total_distance, 2)
        ws.cell(row=total_row, column=2).number_format = '#,##0.00'
        ws.cell(row=total_row, column=3).value = round(total_driving_minutes / 60, 1)
        ws.cell(row=total_row, column=3).number_format = '0.0'
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = self.total_fill
            cell.font = Font(name='맑은 고딕', size=10, bold=True)
            cell.border = self.border
            cell.alignment = self.data_alignment_center
        
        self._auto_adjust_column_width(ws)
        
        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Custom report generated: {len(records)} records")
        return output.getvalue()
