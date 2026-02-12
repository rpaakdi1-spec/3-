"""
재무 대시보드 보고서 내보내기 서비스
Excel 및 PDF 형식으로 재무 데이터를 생성합니다.
"""
from typing import Dict, Any, List, Optional
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)


class FinancialReportExporter:
    """재무 대시보드 보고서 내보내기"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # ============= Excel 생성 =============
    
    def generate_excel(
        self,
        summary_data: Dict[str, Any],
        monthly_trends: List[Dict[str, Any]],
        top_clients: List[Dict[str, Any]],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> BytesIO:
        """
        재무 대시보드 데이터를 Excel 파일로 생성
        
        Args:
            summary_data: 재무 요약 데이터 (4개 카드)
            monthly_trends: 월별 추이 데이터
            top_clients: TOP 10 거래처 데이터
            start_date: 조회 시작일
            end_date: 조회 종료일
            
        Returns:
            BytesIO: Excel 파일 바이너리 데이터
        """
        wb = Workbook()
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 1. 요약 시트
        self._create_summary_sheet(wb, summary_data, start_date, end_date)
        
        # 2. 월별 추이 시트
        self._create_trends_sheet(wb, monthly_trends)
        
        # 3. TOP 10 거래처 시트
        self._create_top_clients_sheet(wb, top_clients)
        
        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Excel 파일 생성 완료")
        return output
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        data: Dict[str, Any],
        start_date: Optional[date],
        end_date: Optional[date]
    ):
        """요약 시트 생성"""
        ws = wb.create_sheet("재무 요약")
        
        # 제목
        ws['A1'] = '재무 대시보드 요약'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # 조회 기간
        period_text = f"조회 기간: {start_date or '전체'} ~ {end_date or '오늘'}"
        ws['A2'] = period_text
        ws.merge_cells('A2:D2')
        
        ws['A3'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A3:D3')
        
        # 헤더
        row = 5
        headers = ['항목', '금액 (₩)', '비율 (%)', '건수']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터
        total_revenue = float(data.get('total_revenue', 0))
        collected_amount = float(data.get('collected_amount', 0))
        collection_rate = float(data.get('collection_rate', 0))
        overdue_amount = float(data.get('overdue_amount', 0))
        overdue_count = int(data.get('overdue_count', 0))
        pending_settlement = float(data.get('pending_settlement', 0))
        receivable_amount = total_revenue - collected_amount
        
        rows_data = [
            ('총 매출액', total_revenue, 100.0, ''),
            ('수금액', collected_amount, collection_rate, ''),
            ('미수금', receivable_amount, 100 - collection_rate, ''),
            ('연체액', overdue_amount, '', overdue_count),
            ('정산 대기', pending_settlement, '', ''),
        ]
        
        for idx, (item, amount, rate, count) in enumerate(rows_data, start=row + 1):
            ws.cell(row=idx, column=1, value=item).border = self.border
            ws.cell(row=idx, column=2, value=amount).border = self.border
            ws.cell(row=idx, column=2).number_format = '#,##0'
            
            if rate:
                ws.cell(row=idx, column=3, value=rate).border = self.border
                ws.cell(row=idx, column=3).number_format = '0.00'
            else:
                ws.cell(row=idx, column=3, value='').border = self.border
            
            ws.cell(row=idx, column=4, value=count).border = self.border
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_trends_sheet(self, wb: Workbook, trends: List[Dict[str, Any]]):
        """월별 추이 시트 생성"""
        ws = wb.create_sheet("월별 추이")
        
        # 제목
        ws['A1'] = '월별 매출/수금 추이'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # 헤더
        row = 3
        headers = ['년월', '매출액 (₩)', '수금액 (₩)', '수금률 (%)', '순이익 (₩)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터
        if not trends:
            ws.cell(row=row + 1, column=1, value='데이터 없음')
        else:
            for idx, trend in enumerate(trends, start=row + 1):
                ws.cell(row=idx, column=1, value=trend.get('month', '')).border = self.border
                ws.cell(row=idx, column=2, value=float(trend.get('revenue', 0))).border = self.border
                ws.cell(row=idx, column=2).number_format = '#,##0'
                ws.cell(row=idx, column=3, value=float(trend.get('collected', 0))).border = self.border
                ws.cell(row=idx, column=3).number_format = '#,##0'
                
                revenue = float(trend.get('revenue', 0))
                collected = float(trend.get('collected', 0))
                collection_rate = (collected / revenue * 100) if revenue > 0 else 0
                ws.cell(row=idx, column=4, value=collection_rate).border = self.border
                ws.cell(row=idx, column=4).number_format = '0.00'
                
                profit = float(trend.get('profit', 0))
                ws.cell(row=idx, column=5, value=profit).border = self.border
                ws.cell(row=idx, column=5).number_format = '#,##0'
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        
        # 차트 추가 (데이터가 있을 경우)
        if trends and len(trends) > 0:
            self._add_line_chart(ws, len(trends), row)
    
    def _create_top_clients_sheet(self, wb: Workbook, clients: List[Dict[str, Any]]):
        """TOP 10 거래처 시트 생성"""
        ws = wb.create_sheet("TOP 10 거래처")
        
        # 제목
        ws['A1'] = 'TOP 10 주요 거래처'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # 헤더
        row = 3
        headers = ['순위', '거래처명', '매출액 (₩)', '청구서 수', '수금률 (%)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 데이터
        if not clients:
            ws.cell(row=row + 1, column=1, value='데이터 없음')
        else:
            for idx, client in enumerate(clients, start=1):
                data_row = row + idx
                ws.cell(row=data_row, column=1, value=idx).border = self.border
                ws.cell(row=data_row, column=2, value=client.get('client_name', '')).border = self.border
                ws.cell(row=data_row, column=3, value=float(client.get('total_revenue', 0))).border = self.border
                ws.cell(row=data_row, column=3).number_format = '#,##0'
                ws.cell(row=data_row, column=4, value=int(client.get('invoice_count', 0))).border = self.border
                ws.cell(row=data_row, column=5, value=float(client.get('collection_rate', 0))).border = self.border
                ws.cell(row=data_row, column=5).number_format = '0.00'
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _add_line_chart(self, ws, data_count: int, start_row: int):
        """월별 추이 라인 차트 추가"""
        chart = LineChart()
        chart.title = "월별 매출/수금 추이"
        chart.style = 10
        chart.y_axis.title = '금액 (₩)'
        chart.x_axis.title = '월'
        
        # 데이터 범위
        data_revenue = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + data_count)
        data_collected = Reference(ws, min_col=3, min_row=start_row, max_row=start_row + data_count)
        categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + data_count)
        
        chart.add_data(data_revenue, titles_from_data=True)
        chart.add_data(data_collected, titles_from_data=True)
        chart.set_categories(categories)
        
        # 차트 위치
        ws.add_chart(chart, f"G3")
    
    # ============= PDF 생성 =============
    
    def generate_pdf(
        self,
        summary_data: Dict[str, Any],
        monthly_trends: List[Dict[str, Any]],
        top_clients: List[Dict[str, Any]],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> BytesIO:
        """
        재무 대시보드 데이터를 PDF 파일로 생성
        
        Args:
            summary_data: 재무 요약 데이터
            monthly_trends: 월별 추이 데이터
            top_clients: TOP 10 거래처 데이터
            start_date: 조회 시작일
            end_date: 조회 종료일
            
        Returns:
            BytesIO: PDF 파일 바이너리 데이터
        """
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # 스타일 정의
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1  # 중앙 정렬
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c5aa0'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # 컨텐츠 생성
        story = []
        
        # 제목
        title = Paragraph("재무 대시보드 보고서", title_style)
        story.append(title)
        
        # 조회 기간
        period_text = f"조회 기간: {start_date or '전체'} ~ {end_date or '오늘'}"
        generation_time = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Paragraph(generation_time, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # 1. 재무 요약
        story.append(Paragraph("1. 재무 요약", heading_style))
        summary_table = self._create_summary_table(summary_data)
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # 2. 월별 추이
        story.append(Paragraph("2. 월별 매출/수금 추이", heading_style))
        if monthly_trends:
            trends_table = self._create_trends_table(monthly_trends)
            story.append(trends_table)
        else:
            story.append(Paragraph("데이터 없음", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # 3. TOP 10 거래처
        story.append(Paragraph("3. TOP 10 주요 거래처", heading_style))
        if top_clients:
            clients_table = self._create_clients_table(top_clients)
            story.append(clients_table)
        else:
            story.append(Paragraph("데이터 없음", styles['Normal']))
        
        # PDF 생성
        doc.build(story)
        output.seek(0)
        
        logger.info("PDF 파일 생성 완료")
        return output
    
    def _create_summary_table(self, data: Dict[str, Any]) -> Table:
        """재무 요약 테이블 생성"""
        total_revenue = float(data.get('total_revenue', 0))
        collected_amount = float(data.get('collected_amount', 0))
        collection_rate = float(data.get('collection_rate', 0))
        overdue_amount = float(data.get('overdue_amount', 0))
        overdue_count = int(data.get('overdue_count', 0))
        pending_settlement = float(data.get('pending_settlement', 0))
        receivable_amount = total_revenue - collected_amount
        
        table_data = [
            ['항목', '금액 (₩)', '비율 (%)', '건수'],
            ['총 매출액', f'{total_revenue:,.0f}', '100.0', ''],
            ['수금액', f'{collected_amount:,.0f}', f'{collection_rate:.2f}', ''],
            ['미수금', f'{receivable_amount:,.0f}', f'{100 - collection_rate:.2f}', ''],
            ['연체액', f'{overdue_amount:,.0f}', '', f'{overdue_count}'],
            ['정산 대기', f'{pending_settlement:,.0f}', '', ''],
        ]
        
        table = Table(table_data, colWidths=[150, 150, 100, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        return table
    
    def _create_trends_table(self, trends: List[Dict[str, Any]]) -> Table:
        """월별 추이 테이블 생성"""
        table_data = [['년월', '매출액 (₩)', '수금액 (₩)', '수금률 (%)', '순이익 (₩)']]
        
        for trend in trends:
            revenue = float(trend.get('revenue', 0))
            collected = float(trend.get('collected', 0))
            collection_rate = (collected / revenue * 100) if revenue > 0 else 0
            profit = float(trend.get('profit', 0))
            
            table_data.append([
                trend.get('month', ''),
                f'{revenue:,.0f}',
                f'{collected:,.0f}',
                f'{collection_rate:.2f}',
                f'{profit:,.0f}'
            ])
        
        table = Table(table_data, colWidths=[100, 130, 130, 100, 130])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        return table
    
    def _create_clients_table(self, clients: List[Dict[str, Any]]) -> Table:
        """TOP 10 거래처 테이블 생성"""
        table_data = [['순위', '거래처명', '매출액 (₩)', '청구서 수', '수금률 (%)']]
        
        for idx, client in enumerate(clients, start=1):
            table_data.append([
                str(idx),
                client.get('client_name', ''),
                f"{float(client.get('total_revenue', 0)):,.0f}",
                str(int(client.get('invoice_count', 0))),
                f"{float(client.get('collection_rate', 0)):.2f}"
            ])
        
        table = Table(table_data, colWidths=[60, 200, 150, 100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        return table


# 전역 인스턴스
financial_report_exporter = FinancialReportExporter()
